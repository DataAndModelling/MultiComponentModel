import openpyxl
import csv
from logger import get_logger
from classes import Blueprint, Part_attributes

logger = get_logger()

def load_part_attributes(filename='parts_data.xlsx'):
    """
    Load part attributes from an Excel file and return a dictionary of Part_attributes instances.
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    parts = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        part_type, failure_hours, life_limit, oh_limit, shape_factor, cost, depot_tat, placeholder = row
        # Apply defaults
        object_name = part_type + "_blueprint"

        failure_hours = float('inf') if failure_hours in (None, '') else failure_hours
        life_limit = float('inf') if life_limit in (None, '') else life_limit
        oh_limit = float('inf') if oh_limit in (None, '') else oh_limit
        cost = 0 if cost in (None, '') else cost
        depot_tat = 0 if depot_tat in (None, '') else depot_tat

        part_object = Part_attributes(
            name=part_type,
            failure_hours=failure_hours,
            life_limit=life_limit,
            shape_factor=shape_factor,
            cost=cost,
            depot_overhaul=oh_limit,
            depot_tat=depot_tat,
            placeholder=placeholder
        )

        parts[object_name] = part_object
        #print("i created a new part!")
        #print(object_name)

    logger.info(f"Loaded {len(parts)} part attributes from '{filename}'")
    #print("after load:", len(Part_attributes.master_list))
    return parts

def load_part_attributes_csv(filename='parts_data.csv'):
    """
    Load part attributes from a CSV file and return a dictionary of Part_attributes instances.
    """
    parts = {}

    with open(filename, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)

        for row in reader:
            part_type = row.get('part_type')
            failure_hours = row.get('failure_hours')
            life_limit = row.get('life_limit')
            oh_limit = row.get('oh_limit')
            shape_factor = row.get('shape_factor')
            cost = row.get('cost')
            depot_tat = row.get('depot_tat')
            placeholder = row.get('placeholder')

            # Apply defaults and convert values
            object_name = part_type + "_blueprint"

            failure_hours = float('inf') if failure_hours in (None, '', 'inf') else float(failure_hours)
            life_limit = float('inf') if life_limit in (None, '', 'inf') else float(life_limit)
            oh_limit = float('inf') if oh_limit in (None, '', 'inf') else float(oh_limit)
            shape_factor = float(shape_factor) if shape_factor not in (None, '') else 1.0
            cost = float(cost) if cost not in (None, '') else 0.0
            depot_tat = float(depot_tat) if depot_tat not in (None, '') else 0.0

            part_object = Part_attributes(
                name=part_type,
                failure_hours=failure_hours,
                life_limit=life_limit,
                shape_factor=shape_factor,
                cost=cost,
                depot_overhaul=oh_limit,
                depot_tat=depot_tat,
                placeholder=placeholder
            )

            parts[object_name] = part_object

    logger.info(f"Loaded {len(parts)} part attributes from '{filename}'")
    return parts

def load_blueprints(filename='blueprint_data.xlsx'):
    """
    Load blueprint hierarchy from Excel and return the root Blueprint object.
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    rows = [(place, part_type, parent_place) for place, part_type, parent_place in sheet.iter_rows(min_row=2, values_only=True)]
    place_to_blueprint = {}
    root = None

    # Identify root node
    for place, part_type, parent_place in rows:
        if parent_place == 'None':
            root = Blueprint(place, part_type)
            place_to_blueprint[place] = root
            break

    if not root:
        logger.error("No root node found where Parent_Place is 'None'")
        return None

    added_parts = {root.place}

    while len(added_parts) < len(rows):
        for place, part_type, parent_place in rows:
            if place not in added_parts and parent_place in place_to_blueprint:
                parent = place_to_blueprint[parent_place]
                parent.add_child(place, part_type)
                place_to_blueprint[place] = parent.children[-1]
                added_parts.add(place)

    logger.info(f"Constructed blueprint tree with {len(added_parts)} nodes.")
    return root

def load_blueprints_csv(filename='blueprint_data.csv'):
    """
    Load blueprint hierarchy from a CSV file and return the root Blueprint object.
    """
    rows = []
    with open(filename, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            place = row.get('place')
            part_type = row.get('part_type')
            parent_place = row.get('parent_place')
            rows.append((place, part_type, parent_place))

    place_to_blueprint = {}
    root = None

    # Identify root node
    for place, part_type, parent_place in rows:
        if parent_place == 'None' or parent_place is None:
            root = Blueprint(place, part_type)
            place_to_blueprint[place] = root
            break

    if not root:
        logger.error("No root node found where parent_place is 'None'")
        return None

    added_parts = {root.place}

    while len(added_parts) < len(rows):
        for place, part_type, parent_place in rows:
            if place not in added_parts and parent_place in place_to_blueprint:
                parent = place_to_blueprint[parent_place]
                parent.add_child(place, part_type)
                place_to_blueprint[place] = parent.children[-1]
                added_parts.add(place)

    logger.info(f"Constructed blueprint tree with {len(added_parts)} nodes.")
    return root

"""

def get_blueprint(filename = 'parts_data.xlsx'):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    return sheet

def load_blueprints_from_excel(filename2):


    # Now, load the blueprint data from the second Excel file (filename2)
    wb = openpyxl.load_workbook(filename2)
    sheet = wb.active

    # List to store rows for easy iteration
    rows = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        place, part_type, parent_place = row
        rows.append((place, part_type, parent_place))

    # Dictionary to store place->blueprint mappings for easy reference
    place_to_blueprint = {}

    # Step 1: Find the root node (where Parent_Place is None)
    root_blueprint = None
    for row in rows:
        place, part_type, parent_place = row
        if parent_place == 'None':
            # This part is the root node
            root_blueprint = Blueprint(place, part_type)
            place_to_blueprint[place] = root_blueprint
            break

    if not root_blueprint:
        logger.info("Error: No root node found (Parent_Place is None).")
        return None

    # Step 2: Iterate through the rows and add children to their respective parents
    added_parts = set([root_blueprint.place])  # Set to track added parts

    # Continue iterating until all parts are added
    while len(added_parts) < len(rows):
        for row in rows:
            place, part_type, parent_place = row
            if place not in added_parts:  # If part is not added yet
                # Check if the parent place is already in place_to_blueprint
                if parent_place in place_to_blueprint:
                    # Add the part as a child of its parent
                    parent_blueprint = place_to_blueprint[parent_place]
                    parent_blueprint.add_child(place, part_type)

                    # Store the part's blueprint reference
                    place_to_blueprint[place] = parent_blueprint.children[-1]
                    added_parts.add(place)

        # If no new parts are added in this iteration, break the loop
        if len(added_parts) == len(rows):
            break

    return root_blueprint
sheet = get_blueprint()

for row in sheet.iter_rows(min_row=2, values_only=True):
    part_type, failure_hours, life_limit, oh_limit, shape_factor, cost, depot_tat,placeholder = row

    # Substitute empty failure_hours or life_limit with float('inf')
    if failure_hours is None or failure_hours == "":
        failure_hours = float('inf')
    if life_limit is None or life_limit == "":
        life_limit = float('inf')
    if oh_limit is None or oh_limit == "":
        oh_limit= float('inf')
    if cost is None or cost == "":
        cost = 0
    if depot_tat is None or depot_tat == "":
        depot_tat = 0

    object_name = part_type + "_blueprint"

    globals()[object_name] = Part_attributes(
        name=part_type,
        failure_hours=failure_hours,
        life_limit=life_limit,
        shape_factor=shape_factor,
        cost=cost,
        depot_overhaul=oh_limit,
        depot_tat=depot_tat,
        placeholder=placeholder
        )


car_blueprint = load_blueprints_from_excel('blueprint_data.xlsx')
"""